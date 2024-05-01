import json
import os
from datetime import datetime, timedelta
from typing import Type

import holidays
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import sessionmaker

from src.projector_backend.dto.abwesenheiten import AbwesenheitDTO, AbwesenheitDetailsDTO, EmployeeDTO, \
    AbwesenheitsRangeDTO
from src.projector_backend.dto.calendar_data import CalenderData
from src.projector_backend.dto.returners import DbResult
from src.projector_backend.entities.abwesenheit_db import AbwesenheitDetails, Employee
from src.projector_backend.excel.excelhelper import ExcelHelper
from src.projector_backend.helpers import data_helper, date_helper


class CalendarService:
    _instance = None

    def __new__(cls, engine):
        if cls._instance is None:
            cls._instance = super(CalendarService, cls).__new__(cls)
            cls._instance.engine = engine
            cls._instance.eh = ExcelHelper()
        return cls._instance

    @classmethod
    def getInstance(cls: Type['CalendarService']) -> 'CalendarService':
        if cls._instance is None:
            raise ValueError("Die Singleton-Instanz wurde noch nicht erstellt.")
        return cls._instance

    def get_calender_data(self, json_format: bool = True) -> CalenderData or str:
        abwesenheitenDTOs: [AbwesenheitDTO] = self._get_all_abwesenheiten(False)
        cd = CalenderData(abwesenheitenDTOs)

        if (json_format):
            return json.dumps(cd, default=data_helper.serialize)
        else:
            return cd

    def get_abwesenheiten_for_psnr(self, personalnummer: int, json_format: bool = True) -> AbwesenheitDTO or str:
        Session = sessionmaker(bind=self.engine)

        with Session() as session:

            employee: Employee = session.query(Employee).filter(Employee.personalnummer == personalnummer).first()


            # TODO: kann man das jetzt nicht vereinfachen?
            subquery = (
                session.query(func.max(AbwesenheitDetails.uploadDatum))
                .filter(AbwesenheitDetails.employee_id == employee.id)
                .subquery()
            )

            abwesenheit_details = (
                session.query(AbwesenheitDetails)
                .filter(AbwesenheitDetails.employee_id == employee.id)
                .filter(AbwesenheitDetails.uploadDatum.in_(session.query(subquery))).all()
            )

            if abwesenheit_details is None:
                abwesenheit_dto = None
            else:
                abwesenheit_dto = AbwesenheitDTO.create_from_db(employee, abwesenheit_details)

            if (json_format):
                return json.dumps(abwesenheit_dto, default=data_helper.serialize)
            else:
                return abwesenheit_dto

    # def add_abwesenheit(self, abwesenheit_dto: AbwesenheitDTO or str):
    #
    #     datum = datetime.now()
    #     dto = abwesenheit_dto
    #     if type(abwesenheit_dto) == str:
    #         jsontext = json.loads(abwesenheit_dto)
    #         dto = AbwesenheitDTO(**jsontext)
    #
    #
    #
    #     session = sessionmaker(bind=self.engine)
    #     with session() as session:
    #         employee = session.query(Employee).filter(Employee.personalnummer).first()
    #
    #         if (not employee):
    #             employee = Employee(abwesenheit_dto.employee_dto.name, abwesenheit_dto.employee_dto.personalnummer)
    #             session.add(employee)
    #             session.refresh(employee)
    #
    #         details = []
    #         detail: AbwesenheitDetailsDTO
    #         for detail in dto.abwesenheitDetails:
    #             if (type(detail) == dict):
    #                 details.append(AbwesenheitDetails(detail["datum"], detail["typ"], uploadDatum=datum))
    #             else:
    #                 details.append(AbwesenheitDetails(detail.datum, detail.typ, uploadDatum=datum))
    #
    #         abw = Employee(dto.name, dto.personalnummer, details, uploadDatum=datum)
    #
    #         session.add(abw)
    #         session.commit()

    def add_abwesenheiten(self, abwesenheiten_dtos: [AbwesenheitDTO] or AbwesenheitDTO or str):

        abwesenheit_details: [AbwesenheitDetails] = []
        datum = datetime.now()

        if type(abwesenheiten_dtos) == str:
            jsontext = json.loads(abwesenheiten_dtos)
            dto = AbwesenheitDTO(**jsontext)
            abwesenheiten_dtos = []
            abwesenheiten_dtos.append(dto)
        elif type(abwesenheiten_dtos) == AbwesenheitDTO:
            dto = abwesenheiten_dtos
            abwesenheiten_dtos = []
            abwesenheiten_dtos.append(dto)

        session = sessionmaker(bind=self.engine)
        with session() as session:
            try:

                dto: AbwesenheitDTO
                for dto in abwesenheiten_dtos:
                    employee_dto = dto.employee
                    employee = session.query(Employee).filter(Employee.personalnummer == employee_dto.personalnummer).first()

                    if (not employee):
                        employee = Employee(employee_dto.name, employee_dto.personalnummer)
                        session.add(employee)
                        session.refresh(employee)


                    detail: AbwesenheitDetailsDTO
                    for detail in dto.abwesenheitDetails:
                        if (type(detail) == dict):
                            abwesenheit_details.append(AbwesenheitDetails(employee, detail["datum"], detail["typ"], uploadDatum=datum))
                        else:
                            abwesenheit_details.append(AbwesenheitDetails(employee,detail.datum, detail.typ, uploadDatum=datum))



                # Füge alle Buchungen hinzu
                session.add_all(abwesenheit_details)

                # Führe die Transaktion durch
                session.commit()

            except IntegrityError as e:
                # Behandle den Fehler speziell für Integritätsverletzungen
                session.rollback()
                print(f"Fehler während der Transaktion: {e}")
                return DbResult(False, e)
            except Exception as e:
                session.rollback()
                print(f"Fehler während der Transaktion: {e}")
                return DbResult(False, e)

        return DbResult(True, "All vacation entries haven been stored successfully.")

    def add_abwesenheits_range(self, abw_r_dto: AbwesenheitsRangeDTO):

        # 1 Alle Tage auflisten
        startdatum = date_helper.from_string_to_date_without_time(abw_r_dto.abwStart)
        enddatum = date_helper.from_string_to_date_without_time(abw_r_dto.abwEnde)

        def date_range(start_date, end_date):
            days = (end_date - start_date).days + 1
            return [start_date + timedelta(days=i) for i in range(days)]

        date_list = date_range(startdatum, enddatum)

        abw_details = []

        try:
            Session = sessionmaker(bind=self.engine)
            with Session() as session:
                employee = session.query(Employee).filter(Employee.personalnummer == abw_r_dto.personalnummer).first()

                for datum in date_list:
                    abw_detail = AbwesenheitDetails(employee, date_helper.from_date_to_string(datum),
                                                    abw_r_dto.abwType)
                    abw_details.append(abw_detail)

                session.add_all(abw_details)
                session.commit()
            return True
        except:
            return False

    def _get_all_abwesenheiten(self, json_format: bool = True):
        Session = sessionmaker(bind=self.engine)

        abw_dtos: [AbwesenheitDTO] = []
        with Session() as session:

            employees = session.query(Employee).all()

            for employee in employees:
                abwesenheitDetailDTOs: [AbwesenheitDetailsDTO] = []
                subquery = (
                    session.query(func.max(AbwesenheitDetails.uploadDatum))
                    .subquery()
                )

                abwesenheit_details = (
                    session.query(AbwesenheitDetails).filter(AbwesenheitDetails.employee_id == employee.id).filter(
                        AbwesenheitDetails.uploadDatum.in_(subquery))
                )

                a: AbwesenheitDetails
                for a in abwesenheit_details:
                    abwesenheitDetailDTOs.append(AbwesenheitDetailsDTO(a.datum, a.typ))

                abw_dtos.append(AbwesenheitDTO(EmployeeDTO.create_from_db(employee), abwesenheitDetailDTOs))

        if json_format:
            return json.dumps(abw_dtos, default=data_helper.serialize)
        else:
            return abw_dtos

    def proceed_demodaten(self, abwesenheiten: [Employee]):
        Session = sessionmaker(bind=self.engine)
        with Session() as session:
            session.query(Employee).delete()
            session.query(AbwesenheitDetails).delete()
            session.commit()

            for abw in abwesenheiten:
                session.add(abw)
            session.commit()

    def get_employees(self, json_format: bool = True):
        Session = sessionmaker(bind=self.engine)

        with Session() as session:

            abwesenheiten = session.query(Employee).all()

            emp_dtos = []
            for abw in abwesenheiten:
                employee_dto = EmployeeDTO.create_from_db(abw)
                emp_dtos.append(employee_dto)

            if (json_format):
                return json.dumps(emp_dtos, default=data_helper.serialize)
            else:
                return emp_dtos

    def prozeed_upload_abwesenheiten(self, filename: str):

        class Monatsabstufung:
            def __init__(self, monat, spalte_start) -> None:
                self.monat: datetime = monat
                self.spalte_beginn = spalte_start
                self.spalte_ende = spalte_start + 1

        try:
            file_path = os.path.join(os.getcwd(), 'uploads', filename)
            wb = self.eh.load_workbook(file_path)
            wb.active: Worksheet = 0

            liste_monate: [Monatsabstufung] = []
            liste_spaltennummern = {}

            # Schneller Check, ob es überhaupt die richtige Datei ist
            name_cell = wb.active['A2'].value
            project_cell = wb.active['B2'].value
            nr_cell = wb.active['C2'].value

            # Grundsätzliches Format überprüfen
            if (name_cell != "Name" or project_cell != "Projekt" or nr_cell != "Perso.-Nr."):
                return DbResult(False,
                                "Die angegebene Datei hat das falsche Format. Handelt es sich dabei wirklich um die Abwesenheitsliste?")

            # 1. Definieren, wo ein Monat anfängt.
            for row in wb.active.iter_rows(values_only=False, min_row=1, max_row=2):

                cell: Cell
                column = 1
                for cell in row:
                    if type(cell.value) == datetime:
                        if len(liste_monate) > 0:
                            liste_monate[-1].spalte_ende = column - 1

                        monat = Monatsabstufung(cell.value, column)
                        liste_monate.append(monat)
                        # print(column, cell.value)
                    column += 1

                liste_monate[-1].spalte_ende = column - 1

            # 2. Definieren, welcher Tag (Zahl) des Monats es ist
            for row in wb.active.iter_rows(values_only=False, min_row=2, max_row=3, min_col=3):

                cell: Cell
                column = 3
                for cell in row:
                    if type(cell.value) == int:
                        liste_spaltennummern[column] = cell.value
                    column += 1

            abw_dtos: [AbwesenheitDTO] = []

            # 3. Urlaube eintragen
            for row in wb.active.iter_rows(values_only=False, min_row=3):

                cell: Cell
                column = 0

                ma_name = row[0].value
                ma_nr = row[2].value

                if (not ma_nr or ma_nr == ""):
                    return DbResult(False,
                                    "Für den Mitarbeiter '" + ma_name + "' wurde keine Personalnummer angegeben.")

                employee_dto = EmployeeDTO(ma_name, ma_nr)

                abwdt_dtos: [AbwesenheitDetailsDTO] = []

                for cell in row:
                    if column < 3:
                        column += 1
                        continue

                    if type(cell.value) == str:

                        cv = cell.value.strip().lower()
                        if cv == "x" or cv == "u" or cv == "k" or cv == "a":
                            tag_zahl = liste_spaltennummern[column + 1]

                            passender_monat: Monatsabstufung = next(
                                (monat for monat in liste_monate if
                                 monat.spalte_beginn <= column + 1 <= monat.spalte_ende),
                                None)

                            if passender_monat is not None:
                                ut = datetime(month=passender_monat.monat.month, day=tag_zahl, year=2024)
                                if cv == "x" or cv == "u":
                                    abwdt_dtos.append(AbwesenheitDetailsDTO(date_helper.from_date_to_string(ut), "u"))
                                else:
                                    abwdt_dtos.append(AbwesenheitDetailsDTO(date_helper.from_date_to_string(ut), "a"))

                            # else:
                            #     print("Kein passendes Element gefunden:")
                            #     print("Passender Monat ist None. ")
                            #     print("Zelle:", cv, cell.col_idx, cell.coordinate)

                    column += 1

                abw_dtos.append(AbwesenheitDTO(employee_dto, abwdt_dtos))

            # 4. Wochenende und Feiertage entnehmen

            deutsche_feiertage = holidays.Germany()

            abw_dto: AbwesenheitDTO
            for abw_dto in abw_dtos:
                abwdtl_dto: AbwesenheitDetailsDTO
                for abwdtl_dto in abw_dto.abwesenheitDetails:

                    converted_date = date_helper.from_string_to_date_without_time(abwdtl_dto.datum)

                    if converted_date in deutsche_feiertage or converted_date.weekday() == 5 or converted_date.weekday() == 6:
                        abw_dto.abwesenheitDetails.remove(abwdtl_dto)

            # 5 in der Datenbank speichern
            result: DbResult = self.add_abwesenheiten(abw_dtos)
            return result

        except IndexError:
            return DbResult(False,
                            "Die Abwesenheitsdatei konnte nicht verarbeitet werden, da das Format nicht korrekt ist. Wurde die richtige Datei ausgewählt?")

        except Exception as e:
            print("### ABW ERROR: ", e)
            return DbResult(False, "Die Abwesenheitsdatei konnte nicht verarbeitet werden. ")

    # def get_employee_by_psnr(self, session, personalnummer):
    #
    #     employee: Employee = session.query(Employee).filter(Employee.personalnummer == personalnummer).first()
    #     if not employee:
    #         Em


