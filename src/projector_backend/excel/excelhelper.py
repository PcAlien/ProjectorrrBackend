from openpyxl import *
from openpyxl.cell import Cell
from openpyxl.styles import *
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHelper:

    def __init__(self) -> None:
        super().__init__()
        self.numbers = numbers

    def load_workbook(self, source):
        return load_workbook(source)

    def get_worksheet(self, source: str, sheetNumber: int) -> Worksheet:
        wb = load_workbook(source)
        wb.active = sheetNumber
        return wb.active

    def get_worksheet_by_name(self, source: str, sheet_name: str) -> Worksheet:
        wb = load_workbook(source)
        wb.active = wb[sheet_name]
        return wb.active

    def create_workbook(self, destination: str, firstSheetName: str) -> Workbook:
        wb = Workbook()
        wb.active.title = firstSheetName
        wb.save(destination)
        return wb

    def copy_sheets_between_files(self, sourceSheet: Worksheet, targetSheet: Worksheet):
        for row in sourceSheet.iter_rows(values_only=True):
            targetSheet.append(row)

    def activate_filter_in_sheet(self, workbook: Workbook, sheets: list = (), allSheets=False):
        if (allSheets):
            for ws in workbook._sheets:
                ws.auto_filter.ref = ws.dimensions
        else:
            for i in sheets:
                for counter, ws in enumerate(workbook._sheets):
                    if (counter == i):
                        ws.auto_filter.ref = ws.dimensions

    def format_column(self, sheet: Worksheet, columnNumber: int, format: str = "dd.mm.YYYY"):
        for x in sheet.iter_cols(min_col=columnNumber + 1, max_col=columnNumber + 1, values_only=False):
            cell: Cell
            for cell in x:
                cell.number_format = format

    def format_first_line_of_every_sheet(self, wb: Workbook):
        font = Font(name="Arial", size=10, color="FFFFFF", bold=True)

        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
                    cell.font = font

            cd = sheet.column_dimensions
            for col in cd:
                cd[col].width = 25

        liste = list(each for each in range(0, len(wb.worksheets)))
        self.activate_filter_in_sheet(workbook=wb, sheets=liste)

    def autosize_current_only_way(self, activeSheet: Worksheet):
        for idx, col in enumerate(activeSheet.columns, 1):
            activeSheet.column_dimensions[get_column_letter(idx)].bestFit = True
