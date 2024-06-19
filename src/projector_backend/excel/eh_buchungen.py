import datetime
from types import NoneType

from openpyxl.cell import Cell
from openpyxl.styles import numbers
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.projector_backend.dto.abwesenheiten import EmployeeDTO
from src.projector_backend.dto.booking_dto import BookingDTO
from src.projector_backend.dto.ma_bookings_summary_dto import MaBookingsSummaryDTO, MaBookingsSummaryElementDTO
from src.projector_backend.dto.monatsaufteilung_dto import MonatsaufteilungDTO, MonatsaufteilungSummaryDTO
from src.projector_backend.entities.ImportFileColumns import ImportFileColumns
from src.projector_backend.excel.excelhelper import ExcelHelper
from src.projector_backend.services.db_service import DBService


class EhBuchungen(ExcelHelper):

    def get_export_type(self, wb: Workbook) -> ImportFileColumns:
        ws: Worksheet = wb.worksheets[0]
        if ws.title == "Loader":
            return DBService.getInstance().get_import_settings(2)
        else:
            return DBService.getInstance().get_import_settings(1)

    def create_booking_dtos_from_export(self, source_file_path: str) -> [BookingDTO]:

        wb: Workbook = self.load_workbook(source_file_path)
        ifc: ImportFileColumns = self.get_export_type(wb)
        bookingDTOs: [BookingDTO] = []
        uploadDatum = datetime.datetime.now()

        if ifc.type == 1:
            ws: Worksheet = wb.active

            bookingDTOs = self._process_excel_reading(uploadDatum, ws, ifc, bookingDTOs)

        elif ifc.type == 2:
            sheetNames = wb.sheetnames

            for sheetName in sheetNames:
                if sheetName != "Loader":
                    ws: Worksheet = wb[sheetName]
                    bookingDTOs = self._process_excel_reading(uploadDatum, ws, ifc, bookingDTOs)

        return bookingDTOs

    def _process_excel_reading(self, uploadDatum: datetime, ws: Worksheet, ifc: ImportFileColumns,
                               bookingDTOs: [BookingDTO]):

        # import time
        # start = time.time()
        # if ifc.delete_empty_lines:
        #     self._delete_summary_rows(ws)
        # stop = time.time()
        #
        # diff = stop-start
        #print("ZEIT: ", diff)

        for row in ws.iter_rows(values_only=True, min_row=2):
            if not self._check_is_summary_row(row):
                dto = BookingDTO(EmployeeDTO(row[ifc.name],
                    row[ifc.personalnummer]),
                    row[ifc.leistungsdatum],
                    row[ifc.fakturierbar],
                    row[ifc.status],
                    row[ifc.bezeichnung],
                    row[ifc.psp],
                    row[ifc.psp_element],
                    row[ifc.stunden],
                    row[ifc.text],
                    row[ifc.erfasst_am],
                    row[ifc.letzte_aenderung],
                    uploaddatum=uploadDatum
                )
                if dto.erstelltAm == None:
                    dto.erstelltAm = dto.letzteAenderung
                bookingDTOs.append(dto)

        return bookingDTOs

    def _check_is_summary_row(self, row):

        cell: Cell
        if type(row[1]) is NoneType:
            return True
        else:
            if row[1].isspace() or row[1] == "":
                return True

        return False


    def _delete_summary_rows(self, sheet: Worksheet):
        toDelete = []
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cell: Cell
            if type(row[1]) is NoneType:
                toDelete.append(i)
            else:
                if row[1].isspace() or row[1] == "":
                    toDelete.append(i)

        # delete useless rows
        for  rowNumber in sorted(toDelete, reverse=True):
            sheet.delete_rows(rowNumber)


        # TODO: lässt sich das kürzen? und schneller machen? das lädt ja ewig

    def export_buchungen(self, psp, booking_dtos: [BookingDTO], monatsaufteilung_dtos: [MonatsaufteilungDTO]):
        export_file_folder = "./exports/"
        export_date = datetime.datetime.now().strftime("%d.%m.%Y")
        export_file_name = f"Buchungen_{psp}_{export_date}.xlsx"
        first_row = ["Name", "Personalnummer", "Datum", "Berechnungsmotiv", "Bearbeitungsstatus", "Bezeichnung",
                     "PSP", "PSP-Element", "Stunden", "Stundensatz", "Umsatz", "Text", "erstellt am",
                     "letzte Änderung"]

        booking_xlsx = self.create_workbook(export_file_folder + export_file_name, "Alle Buchungen")

        first_sheet = booking_xlsx.active

        self._fill_bookings_in_exportfile(first_sheet, first_row, booking_dtos)

        month_dto: MonatsaufteilungDTO
        for month_dto in monatsaufteilung_dtos:
            active_sheet = booking_xlsx.create_sheet(month_dto.monat)
            self._fill_bookings_in_exportfile(active_sheet, first_row, month_dto.bookings)

        self.format_first_line_of_every_sheet(booking_xlsx)
        booking_xlsx.save(export_file_folder + export_file_name)

        return export_file_name

    def export_umsaetze(self, psp, mab_summary_dto: MaBookingsSummaryDTO,
                        monatsaufteilung_sum_dtos: [MonatsaufteilungSummaryDTO], budget: float):
        export_file_folder = "./exports/"
        export_date = datetime.datetime.now().strftime("%d.%m.%Y")
        export_file_name = f"Umsätze_{psp}_{export_date}.xlsx"
        first_row = ["Name", "Personalnummer", "PSP-Element", "Stunden", "Stundensatz", "Umsatz"]

        booking_xlsx = self.create_workbook(export_file_folder + export_file_name, "Alle Umsätze")

        first_sheet = booking_xlsx.active

        self._fill_umsaetze_in_exportfile(first_sheet, first_row, mab_summary_dto.bookings)

        month_dto: MonatsaufteilungSummaryDTO
        for month_dto in monatsaufteilung_sum_dtos:
            active_sheet = booking_xlsx.create_sheet(month_dto.monat)
            self._fill_umsaetze_in_exportfile(active_sheet, first_row, month_dto.maBookingsSummary.bookings)

        # Weiteres Sheet: Zusammenfassung:
        active_sheet = booking_xlsx.create_sheet("Budgetübersicht")
        active_sheet.append(["Monat", "Gesamtumsatz"])
        for month_dto in monatsaufteilung_sum_dtos:
            monat = month_dto.monat
            umsatz = month_dto.maBookingsSummary.sum
            active_sheet.append([monat, umsatz])

        active_sheet.append(["", ""])
        active_sheet.append(["GESAMTUMSATZ", mab_summary_dto.sum])
        active_sheet.append(["Projektbudget", budget])
        active_sheet.append(["Restbudget", budget - mab_summary_dto.sum])
        self.format_column(active_sheet, 1, '#,##0.00 €')
        self.autosize_current_only_way(active_sheet)

        self.format_first_line_of_every_sheet(booking_xlsx)
        booking_xlsx.save(export_file_folder + export_file_name)

        return export_file_name

    def _fill_bookings_in_exportfile(self, worksheet, first_row: [], booking_dtos: [BookingDTO]):
        worksheet.append(first_row)
        dto: BookingDTO
        for dto in booking_dtos:
            worksheet.append(
                [dto.employee.name,dto.employee.personalnummer, dto.datum, dto.berechnungsmotiv, dto.bearbeitungsstatus, dto.bezeichnung,
                 dto.psp, dto.pspElement, dto.stunden, dto.stundensatz, dto.umsatz, dto.text,
                 dto.erstelltAm, dto.letzteAenderung])

        self.format_column(worksheet, 8, numbers.FORMAT_NUMBER_00)
        self.format_column(worksheet, 9, '#,##0.00 €')
        self.format_column(worksheet, 10, '#,##0.00 €')
        self.autosize_current_only_way(worksheet)

    def _fill_umsaetze_in_exportfile(self, worksheet, first_row: [], mas_dtos: [MaBookingsSummaryElementDTO]):
        worksheet.append(first_row)
        dto: MaBookingsSummaryElementDTO
        for dto in mas_dtos:
            worksheet.append(
                [dto.employee.name, dto.employee.personalnummer, dto.psp_element, dto.stunden, dto.stundensatz, dto.umsatz])

        self.format_column(worksheet, 3, numbers.FORMAT_NUMBER_00)
        self.format_column(worksheet, 4, '#,##0.00 €')
        self.format_column(worksheet, 5, '#,##0.00 €')
        self.autosize_current_only_way(worksheet)
